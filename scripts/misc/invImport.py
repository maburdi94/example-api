################################################################################
#Utility script to read inventory records from Virun's inventory spreadsheet   #
#Requires the openpyxl package and python's mysql connector                    #
#install using: pip3 install mysql-connector-python                            #
#pip3 install openpyxl                                                         #
#Written by Eric Lara (remigar@csu.fullerton.edu)                              #
################################################################################

#some notes
#lot 20004: supplier info not imported due to invalid date (2019 date for a 2020 lot)
#RM0109: Lots not imported because first lot does not have a lot number
#need to also record initial amount received, requires schema change

from openpyxl import load_workbook
import mysql.connector
import datetime
import os
import urllib.request
import string

#attempt to connect to the database
print("Connecting to database...")
try:
	db = mysql.connector.connect(host = "localhost", user = "root", password = "123456", database = "Virun_Inventory", auth_plugin = "mysql_native_password")
	dbCursor = db.cursor()
except:
	print("Database connection failed.")
	quit(1)

fn_inv = 'Internal Lot Code Inventory Sheet.xlsx'
#open the rm receiving spreadsheets
#we want a dictionary of years mapping to their respective spreadsheets
#i.e. recvSheets["2020"] maps to the 2020 receiving sheet
fn_recv = ['Inventory Receiving and Internal Lot Code List (2016).xlsx',
			 'Inventory Receiving and Internal Lot Code List (2017).xlsx',
			 'Inventory Receiving and Internal Lot Code List (2018).xlsx',
			 'Inventory Receiving and Internal Lot Code List (2019).xlsx',
			 'Internal Receiving Raw Materials 2020.xlsx']

#get the inventory spreadsheet if it doesn't already exist in the local directory
url = "https://github.com/maburdi94/virun-api/raw/master/scripts/misc/spreadsheets/{}".format(fn_inv.replace(" ", "%20"))
try:
	urllib.request.urlretrieve(url, fn_inv)
except:
	print("Could not retrieve sheet {}.".format(fn_inv))

#check if each file in fn_recv exists
#if it doesn't exist, download from the repo (https://github.com/maburdi94/virun-api/tree/master/scripts/misc
for fn in fn_recv:
	if not os.path.exists(fn):
		url = "https://github.com/maburdi94/virun-api/raw/master/scripts/misc/spreadsheets/{}".format(fn.replace(" ", "%20"))
		print(url)
		try:
			urllib.request.urlretrieve(url, fn)
		except:
			print("could not retrieve sheet {}.".format(fn))
			exit(1)

recvWBs = {}
recvSheets = {}

#load the inventory spreadsheet, this will take a while
print("Loading Spreadsheet {}...".format(fn_inv))
wb = load_workbook(filename = fn_inv, data_only=True)
#retrieve the production inventory worksheet
prodSheet = wb['PRODUCTION']


for fn in fn_recv:
	year = str().join(filter(str.isdigit, fn))
	print("Loading Spreadsheet {}...".format(fn))
	recvWBs[year] = load_workbook(filename = fn, data_only = True)
	#2020's sheet is titled slightly differently
	if "2020" in fn:
		recvSheets[year] = recvWBs[year]['Internal Receiving Log {}'.format(year)]
	else:
		recvSheets[year] = recvWBs[year]['REC-01 Receiving Log {}'.format(year)]



def getCellValue(sheet, column, row):
	return sheet['{}{}'.format(column, row)].value

def goToNextLotorRM(sheet, curRow, curRM):
	#given a row in a sheet, proceed to the next lot number of a rm number or the next rm number if all lots have been recorded
	newLot = False
	newRM = False
	result = curRow + 1

	while(getCellValue(sheet, 'C', result) == None or "L" not in str(getCellValue(sheet, 'C', result))):
		result+=1
		aVal = str(getCellValue(sheet, 'A', result))
		if aVal != None and "RM" in aVal and aVal != curRM:
			newRM = True
			break
	#result should now store the row value of a lot within the current RM, or a new RM to process
	return (result, newLot, newRM)

def offsetChar(s, offset):
	return chr(ord(s)+offset)


row = 1
while(True):
	recordTemp = []
	#is there a rm record on this row?
	if prodSheet['A{}'.format(row)].value != None and "RM" in str(prodSheet['A{}'.format(row)].value):
		#if yes, retrieve the record RMNumber and Description
		#our db schema defines the rm number as a 6 digit integer
		#thus we strip all non-numeric characters
		rmNum = getCellValue(prodSheet, 'A', row)
		rmNum = int(str().join(filter(str.isdigit, rmNum)))
		recordTemp.append(rmNum)
		rmDesc = getCellValue(prodSheet,'B', row)
		recordTemp.append(rmDesc)

		#insert rmNum and rmDesc into the RawMaterial TABLE
		sqlRM = "INSERT INTO RawMaterial (rm, name, type) VALUES (%s, %s, %s)"
		#from what I can tell, the inventory sheet does not specify a material type
		rmTup = tuple(recordTemp + ['None'])
		try:
			dbCursor.execute(sqlRM, rmTup)
			print ("Inserted RM#{} {} into RawMaterial table.".format(rmNum,rmDesc))
		except:
			print("failed to insert RM#{} {} into RawMaterial table".format(rmNum, rmDesc))
			row+=1
			continue

		#for each RMNumber, make a record for each specific lot of of that RMNumber
		#lots begin on the row after the rmnumber/description
		'''if "L" in str(getCellValue(prodSheet, 'C', row+1)).upper():
			lotrow = row + 1
		else:
			lotrow = goToNextLotorRM(prodSheet, row, rmNum)[0]'''
		#NOTE: if the row immediately after the rm row doesn't have a valid lot record, lots for that rm won't be recorded
		lotrow = row + 1
		while (getCellValue(prodSheet,'C', lotrow) != None):
			record = [rmNum]
			lotNum = getCellValue(prodSheet,'C', lotrow).upper()#note: upper() call was added due to an inconsistency in lots L20096 and L20097
			if "L" in str(lotNum):
				#retrieve date the raw material was received
				rcvDate = getCellValue(prodSheet,'D', lotrow)

				#set the inital balance of the lot, we will later check if there has been any transactions
				#balance = prodSheet['I{}'.format(lotrow)].value
				balance = getCellValue(prodSheet, 'I', lotrow)

				record.append(lotNum)
				record.append(rcvDate)
				#TODO: get the most up-to-date balance
				#iterate through the rows until we find the most up-to-date balance
				#progress through rows while the next row does not have a new lot code or the next row's quantity cell isn't empty
				while (True):
					balance = getCellValue(prodSheet, 'I', lotrow)
					nextQtyCell = getCellValue(prodSheet, 'I', lotrow+1)
					nextLotCell = str(getCellValue(prodSheet, 'C', lotrow+1)).upper()
					nextRMCell  = str(getCellValue(prodSheet, 'A', lotrow+1))
					#if the next row is empty, we have the correct balance
					if nextQtyCell == None:
						break
					#if the next row has a new lot number, we have the correct balance
					elif "L" in nextLotCell and nextLotCell != lotNum:
						break
					#if the next row has a new rm record, we have the correct balance
					elif nextRMCell != None and "RM" in nextRMCell:
						break
					#if none of the above conditions are satisfied, we increment lotrow and continue
					lotrow+=1


				record.append(balance)

				#progress down the rows until we hit a new lot OR a new rm#

			#from this spreadsheet we have extracted rm, lot, receive date, and quantity, but not supplier, rack location, or manufacturer
			#TODO: if we have data available for the year this lot was received, retrieve the supplier information from that Sheet
			temp = ['0001', 'N/A', 'N/A'] #if we cant find the supplier information, we insert the lot with some dummy data
			#get the year the lot was received
			try:
				rcvYear = str(rcvDate.year)
			except:
				rcvYear = "N/A"
			offset16 = 0

			if rcvYear in recvSheets:
				#find the row in the sheet corresponding to the lot number
				#instead of writing slightly different logic for 2016, add an offset to the column we're selecting from if the lot is from 2016
				if rcvYear == "2016":
					offset16 = 3
					rcvrow = 4
				else:
					offset16 = 0
					rcvrow = 2
				while(getCellValue(recvSheets[rcvYear], offsetChar("A", offset16), rcvrow) != None):
					v_LotNum = getCellValue(recvSheets[rcvYear], offsetChar("A", offset16), rcvrow)
					if v_LotNum == lotNum:
						break
					rcvrow += 1
				#retrieve the mfr lot number for this lotTup
				temp[2] = getCellValue(recvSheets[rcvYear], offsetChar("D", offset16), rcvrow)
				print(temp[2])



				#query the database to get the supplier id corresponding to the row's supplier name
				#in an effort to reduce duplicate Supplier entries, we stripped all whitepsace and forced all names uppercase, so we do the same here.
				#also removed punctuation
				suppName = str(getCellValue(recvSheets[rcvYear], offsetChar("C", offset16), rcvrow)).strip().upper()
				suppName = suppName.translate(str.maketrans('','',string.punctuation))
				suppQry = "SELECT id FROM Supplier WHERE name = %s"
				dbCursor.execute(suppQry, (suppName, ))
				qryResult = dbCursor.fetchall()
				temp[0] = qryResult[0][0]


			#DONE...? insert this extracted record into the database
			sqlLot = "INSERT INTO Inventory (rm, lot, arrived, qty, supplier, rack, mfr) VALUES (%s, %s, %s, %s, %s, %s, %s)"
			lotTup = tuple(record + temp)
			print (lotTup)
			try:
				dbCursor.execute(sqlLot,lotTup)
				print("Inserted lot#{} of rm#{} into table Inventory.".format(lotNum, rmNum))
				#move onto next lot, we break if we encounter a new rmNum
				result = goToNextLotorRM(prodSheet, lotrow, rmNum)
				lotrow = result[0]
				if result[2]:
					break
			except:
				print("Failed to insert lot#{} of rm#{}".format(lotNum, rmNum))
				#move onto the next lot
				result = goToNextLotorRM(prodSheet, lotrow, rmNum)
				lotrow = result[0]
				if result[2]:
					break

	if recordTemp != []:
		pass
		#print (recordTemp)
	row+=1
	if row == 25000:
		db.commit()
		db.close()
		break
#TODO: delete the sheet files after we're done?
for fn in fn_recv:
	os.remove(fn)
	print("Removed file {}".format(fn))
os.remove(fn_inv)

print("Done.")
