################################################################################
#utility to import supplier information from Virun's rm receiving spreadsheets #
#current: parse through 2016-2020 rm receiving sheets                          #
#grab all unique supplier names and insert into db                             #
#some issues with duplicates (e.g. "AIC" "aic" "AIC ")                         #
#Written by Eric Lara (remigar@csu.fullerton.edu)                              #
################################################################################

from openpyxl import load_workbook
import mysql.connector
import string
import os.path

#attempt to connect to the database
print("Connecting to database...")
try:
	db = mysql.connector.connect(host = "localhost", user = "root", password = "123456", database = "Virun_Inventory", auth_plugin = "mysql_native_password")
	dbCursor = db.cursor()
except:
	print("Database connection failed.")
	quit(1)
fn_recv = ['Inventory Receiving and Internal Lot Code List (2016).xlsx',
			 'Inventory Receiving and Internal Lot Code List (2017).xlsx',
			 'Inventory Receiving and Internal Lot Code List (2018).xlsx',
			 'Inventory Receiving and Internal Lot Code List (2019).xlsx',
			 'Internal Receiving Raw Materials 2020.xlsx']

#check if each file in fn_recv exists
#if it doesn't exist, download from the repo (https://github.com/maburdi94/virun-api/tree/master/scripts/misc
for fn in fn_recv:
	if not os.path.exists(fn):
		url = "https://github.com/maburdi94/virun-api/raw/master/scripts/misc/{}".format(fn.replace(" ", "%20"))
		print(url)
		try:
			urllib.request.urlretrieve(url, fn)
		except:
			print("could not retrieve sheet {}.".format(fn))
			exit(1)


#fn2020 = 'Internal Receiving Raw Materials 2020.xlsx'

#print("Loading Spreadsheet {}...".format(fn2020))
#wb = load_workbook(filename = fn2020, data_only=True)
#retrieve the production inventory worksheet
#recvSheet = wb['Internal Receiving Log 2020']

recvWBs = {}
recvSheets = {}

for fn in fn_recv:
	year = str().join(filter(str.isdigit, fn))
	print("Loading Spreadsheet {}...".format(fn))
	recvWBs[year] = load_workbook(filename = fn, data_only = True)
	#2020's sheet is titled slightly differently
	if "2020" in fn:
		recvSheets[year] = recvWBs[year]['Internal Receiving Log {}'.format(year)]
	else:
		recvSheets[year] = recvWBs[year]['REC-01 Receiving Log {}'.format(year)]

#we now have a dictionary of inventory worksheets, keyed by year.

def getCellValue(sheet, column, row):
	return sheet['{}{}'.format(column, row)].value

uniqueSuppNames = []
'''while(getCellValue(recvSheet, "A", row) != None):
	#each row has a unique lot of some rm
	rowSupp = getCellValue(recvSheet, "C", row)
	#is this name unique?
	#if yes, add to the list
	if rowSupp not in uniqueSuppNames:
		uniqueSuppNames.append(rowSupp)
	row += 1'''
#2016 requires slightly modified logic due to being formatted a little bit differently
#in the interest of reducing duplicate supplier entries, we strip leading and trailing whitespace from the names we retrieve
# IDEA: store retrieved name as all uppercase as well?
# IDEA: Remove all punctuation?
# IDEA: before inserting check if each record is a substring of some other string in the list; if so, remove it
#Maybe ask for a confirmation first?
row = 4
while(getCellValue(recvSheets["2016"], "D", row) != None):
	#each row has a unique lot of some rm
	rowSupp = str(getCellValue(recvSheets["2016"], "F", row)).strip().upper()
	rowSupp = rowSupp.translate(str.maketrans('','',string.punctuation))
	#is this name unique?
	#if yes, add to the list
	if rowSupp not in uniqueSuppNames and rowSupp != "None":
		uniqueSuppNames.append(rowSupp)
	row += 1

#years 2017 through 2020 should import the same way
for year in range(2017, 2021):
	yearstr = str(year)
	row = 2
	while(getCellValue(recvSheets[yearstr], "A", row) != None):
		#each row has a unique lot of some rm
		rowSupp = str(getCellValue(recvSheets[yearstr], "C", row)).strip().upper()
		#remove punctuation
		rowSupp = rowSupp.translate(str.maketrans('','',string.punctuation))
		#is this name unique?
		#if yes, add to the list
		if rowSupp not in uniqueSuppNames and rowSupp != "NONE":
			uniqueSuppNames.append(rowSupp)
		row += 1

dupeCount = 0
#we should now have a list of all suppliers that have shipped to Virun at this point
#try to filter out potential duplicates by finding that are substrings of another entry
for entry in uniqueSuppNames:
	for entry2 in uniqueSuppNames:
		if entry == entry2:
			continue
		elif entry in entry2:
			print("possible duplicates: {} , {}".format(entry, entry2))
			dupeCount+=1
print("Possible Duplicates Found: {}".format(dupeCount))

print(uniqueSuppNames)
print("Suppliers Retrieved: {}".format(len(uniqueSuppNames)))

sqlSupp = "INSERT INTO Supplier (name, phone, email) VALUES (%s, %s, %s)"
for supp in uniqueSuppNames:
	suppTup = tuple([supp] + ["N/A", "N/A"])
	try:
		dbCursor.execute(sqlSupp, suppTup)
		print ("Inserted Supplier {} into Supplier table.".format(supp))
	except:
		print("failed to insert Supplier {} into Supplier table".format(supp))

db.commit()
db.close()
print("Done.")
